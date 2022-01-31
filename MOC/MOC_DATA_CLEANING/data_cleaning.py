import xlrd
import openpyxl as xl
import re
import xlsxwriter

# Read data from this sheet
workbook = xlrd.open_workbook('to_be_clean.xlsx')
# workbook = load_workbook(filename = "1_33000.xlsx", read_only=True)
worksheet = workbook.sheet_by_name('main_data')
# worksheet = workbook['Sheet1']
# Create new file
outWorkbook = xlsxwriter.Workbook("cleaned_file.xlsx")

# create sheet name called T_business_categories
outSheet_raw_data = outWorkbook.add_worksheet("Raw Data")
outSheet_T_business_category = outWorkbook.add_worksheet("T_business_categories")
outSheet_Master_Company_Code = outWorkbook.add_worksheet("Master_Company_Code")
outSheet_Master_Business_Code = outWorkbook.add_worksheet("Master_Business Code")
outSheet_Master_Main_Business_Activities = outWorkbook.add_worksheet("Master_Main Business Activities")
outSheet_T_CompanyData = outWorkbook.add_worksheet("T_CompanyData")
outSheet_T_Director_Data = outWorkbook.add_worksheet("T_Director_Data")
# Create header for sheet T_business_categories
outSheet_T_business_category.write("A1", "company_ids")
outSheet_T_business_category.write("B1", "business_activities")
outSheet_T_business_category.write("C1", "business_objective")
outSheet_T_business_category.write("D1", "business_main_activities")
outSheet_T_business_category.write("E1", "business_objective_code")
outSheet_T_business_category.write("F1", "business_activities_code")
# Create header for sheet Master_Company_Code
outSheet_Master_Company_Code.write("A1", "Company_ids")
outSheet_Master_Company_Code.write("B1", "master_company_code_name_kh")
outSheet_Master_Company_Code.write("C1", "master_company_code_name_en")
# Create Header for sheet Master Business Code
outSheet_Master_Business_Code.write("A1", "Business Objective Code")
outSheet_Master_Business_Code.write("A1", "Business Objective")
# Create header for sheet Master_Main Business Activities
outSheet_Master_Main_Business_Activities.write("A1", "Main Business Activities Code")
outSheet_Master_Main_Business_Activities.write("B1", "Main Business Activities")
# Create header for sheet T_CompanyData
outSheet_T_CompanyData.write("A1", "Company_ID")
outSheet_T_CompanyData.write("B1", "Company Name (in Khmer)")
outSheet_T_CompanyData.write("C1", "Company Name (in English)")
outSheet_T_CompanyData.write("D1", "Original Entity Identifier")
outSheet_T_CompanyData.write("E1", "Company Status")
outSheet_T_CompanyData.write("F1", "Incorporation Date")
outSheet_T_CompanyData.write("G1", "Re-Registration Date")
outSheet_T_CompanyData.write("H1", "Tax Identification Number (TIN)")
outSheet_T_CompanyData.write("I1", "Tax Registration Date")
outSheet_T_CompanyData.write("J1", "Annual Return Last Filed on")
outSheet_T_CompanyData.write("K1", "Male")
outSheet_T_CompanyData.write("L1", "Female")
outSheet_T_CompanyData.write("M1", "Number of Cambodian Employees")
outSheet_T_CompanyData.write("N1", "Number of Foreign Employees")
outSheet_T_CompanyData.write("O1", "Contact Email")
outSheet_T_CompanyData.write("P1", "Contact Telephone Number")
outSheet_T_CompanyData.write("Q1", "Physical Registered Office Address")
outSheet_T_CompanyData.write("R1", "Start Date")
outSheet_T_CompanyData.write("S1", "Postal Registered Office Address")
outSheet_T_CompanyData.write("T1", "Start Date")
# Create Header for sheet T_Director_Data
outSheet_T_Director_Data.write("A1", "company_ids")
outSheet_T_Director_Data.write("B1", "Directors")
outSheet_T_Director_Data.write("C1", "Name in Khmer")
outSheet_T_Director_Data.write("D1", "Name in English")
outSheet_T_Director_Data.write("E1", "Postal_Registered_Office_Address")
outSheet_T_Director_Data.write("F1", "Telephone")
outSheet_T_Director_Data.write("G1", "Chairman_of_the_Board_of_Directors")

# Define variable for sheet T_business_categories
company_ids = []
business_activities = []
business_objective = []
business_objective_text = []
business_main_activities = []
business_main_activities_text = []
business_objective_code = []
business_activities_code = []

# Define variable for sheet master company code
master_company_code_company_id = []
master_company_code_name_kh = []
master_company_code_name_en = []
# Define variable for sheet T_CompanyData
# company_data
company_data_company_id = []
company_data_name_kh = []
company_data_name_en = []
company_data_oei = []
company_data_company_status = []
company_data_incorporate_date = []
company_data_rrd = []
company_data_tin = []
company_data_trd = []
company_data_arl = []
company_data_male = []
company_data_female = []
company_data_nce = []
company_data_nfe = []
company_data_ce = []
company_data_ctn = []
company_data_proa = []
company_data_sd = []
company_data_proa_last = []
company_data_sd_last = []

# Define variable for sheet T_Director_Data
director_company_ids = []
director_list = []
director_name_kh = []
director_name_en = []
director_post_address = []
director_telephone = []
director_chairman = []


def t_business_category(sheet_column_index, loop_range):
    initial_value = 0
    for i in range(loop_range):
        # len(worksheet.col_values(sheet_column_index, 2))
        for item in range(len(worksheet.col_values(sheet_column_index, 2))):
            value = worksheet.col_values(45 + initial_value, 2)[item]
            if value != " ":
                try:
                    print(str(i) + " running on t_business_category " + str(item))
                    business_objective_value = worksheet.col_values(sheet_column_index + initial_value, 2)[item]
                    business_main_activities_value = worksheet.col_values(sheet_column_index + 1 + initial_value, 2)[
                        item]
                    company_id_value = worksheet.col_values(0, 2)[item]
                    business_objective_code_value = re.findall("\d+", business_objective_value)[0]
                    business_activities_code_value = re.findall("\d+", business_main_activities_value)[0]
                    business_objective_text_value = " ".join(re.findall("[a-zA-Z]+", business_objective_value))
                    business_main_activities_text_text = " ".join(re.findall("[a-zA-Z]+", business_main_activities_value))

                    business_objective_text.append(business_objective_text_value)
                    business_main_activities_text.append(business_main_activities_text_text)
                    business_objective.append(business_objective_value)
                    business_main_activities.append(business_main_activities_value)
                    business_activities.append("Business Activity " + str(i + 1))
                    business_objective_code.append(business_objective_code_value)
                    business_activities_code.append(business_activities_code_value)
                    company_ids.append(company_id_value)
                except:
                    pass

        initial_value = initial_value + 2
    for item in range(len(business_objective)):
        outSheet_T_business_category.write(item + 1, 0, company_ids[item])
        outSheet_T_business_category.write(item + 1, 1, business_activities[item])
        outSheet_T_business_category.write(item + 1, 2, business_objective[item])
        outSheet_T_business_category.write(item + 1, 3, business_main_activities[item])
        outSheet_T_business_category.write(item + 1, 4, business_objective_code[item])
        outSheet_T_business_category.write(item + 1, 5, business_activities_code[item])

    for item in range(len(business_objective_code)):
        outSheet_Master_Business_Code.write(item + 1, 0, business_objective_code[item])
        outSheet_Master_Business_Code.write(item + 1, 1, business_objective_text[item])
    for item in range(len(business_activities_code)):
        outSheet_Master_Main_Business_Activities.write(item + 1, 0, business_activities_code[item])
        outSheet_Master_Main_Business_Activities.write(item + 1, 1, business_main_activities_text[item])
    # outWorkbook.close()


def master_company_code():
    for item in range(len(worksheet.col_values(0, 2))):
        value = worksheet.col_values(0, 2)[item]
        if value != " ":
            try:
                print("running on master_company_code: " + str(item))
                master_company_code_company_id_value = worksheet.col_values(0, 2)[item]
                master_company_code_name_kh_value = worksheet.col_values(1, 2)[item]
                master_company_code_name_en_value = worksheet.col_values(2, 2)[item]
                master_company_code_company_id.append(master_company_code_company_id_value)
                master_company_code_name_kh.append(master_company_code_name_kh_value)
                master_company_code_name_en.append(master_company_code_name_en_value)
            except:
                pass
    for item in range(len(master_company_code_company_id)):
        outSheet_Master_Company_Code.write(item + 1, 0, master_company_code_company_id[item])
        outSheet_Master_Company_Code.write(item + 1, 1, master_company_code_name_kh[item])
        outSheet_Master_Company_Code.write(item + 1, 2, master_company_code_name_en[item])


def company_data():
    for item in range(len(worksheet.col_values(0, 2))):
        value = worksheet.col_values(0, 2)[item]
        if value != " ":
            try:
                print("running on company data: " + str(item))
                company_data_company_id_value = worksheet.col_values(0, 2)[item]
                company_data_name_kh_value = worksheet.col_values(1, 2)[item]
                company_data_name_en_value = worksheet.col_values(2, 2)[item]
                company_data_oei_value = worksheet.col_values(3, 2)[item]
                company_data_company_status_value = worksheet.col_values(4, 2)[item]
                company_data_incorporate_date_value = worksheet.col_values(5, 2)[item]
                company_data_rrd_value = worksheet.col_values(6, 2)[item]
                company_data_tin_value = worksheet.col_values(7, 2)[item]
                company_data_trd_value = worksheet.col_values(8, 2)[item]
                company_data_arl_value = worksheet.col_values(9, 2)[item]
                company_data_male_value = worksheet.col_values(10, 2)[item]
                company_data_female_value = worksheet.col_values(11, 2)[item]
                company_data_nce_value = worksheet.col_values(12, 2)[item]
                company_data_nfe_value = worksheet.col_values(13, 2)[item]
                company_data_ce_value = worksheet.col_values(14, 2)[item]
                company_data_ctn_value = worksheet.col_values(15, 2)[item]
                company_data_proa_value = worksheet.col_values(16, 2)[item]
                company_data_sd_value = worksheet.col_values(17, 2)[item]
                company_data_proa_last_value = worksheet.col_values(18, 2)[item]
                company_data_sd_last_value = worksheet.col_values(19, 2)[item]

                company_data_company_id.append(company_data_company_id_value)
                company_data_name_kh.append(company_data_name_kh_value)
                company_data_name_en.append(company_data_name_en_value)
                company_data_oei.append(company_data_oei_value)
                company_data_company_status.append(company_data_company_status_value)
                company_data_incorporate_date.append(company_data_incorporate_date_value)
                company_data_rrd.append(company_data_rrd_value)
                company_data_tin.append(company_data_tin_value)
                company_data_trd.append(company_data_trd_value)
                company_data_arl.append(company_data_arl_value)
                company_data_male.append(company_data_male_value)
                company_data_female.append(company_data_female_value)
                company_data_nce.append(company_data_nce_value)
                company_data_nfe.append(company_data_nfe_value)
                company_data_ce.append(company_data_ce_value)
                company_data_ctn.append(company_data_ctn_value)
                company_data_proa.append(company_data_proa_value)
                company_data_sd.append(company_data_sd_value)
                company_data_proa_last.append(company_data_proa_last_value)
                company_data_sd_last.append(company_data_sd_last_value)


            except:
                pass
    for item in range(len(company_data_company_id)):
        outSheet_T_CompanyData.write(item + 1, 0, company_data_company_id[item])
        outSheet_T_CompanyData.write(item + 1, 1, company_data_name_kh[item])
        outSheet_T_CompanyData.write(item + 1, 2, company_data_name_en[item])
        outSheet_T_CompanyData.write(item + 1, 3, company_data_oei[item])
        outSheet_T_CompanyData.write(item + 1, 4, company_data_company_status[item])
        outSheet_T_CompanyData.write(item + 1, 5, company_data_incorporate_date[item])
        outSheet_T_CompanyData.write(item + 1, 6, company_data_rrd[item])
        outSheet_T_CompanyData.write(item + 1, 7, company_data_tin[item])
        outSheet_T_CompanyData.write(item + 1, 8, company_data_trd[item])
        outSheet_T_CompanyData.write(item + 1, 9, company_data_arl[item])
        outSheet_T_CompanyData.write(item + 1, 10, company_data_male[item])
        outSheet_T_CompanyData.write(item + 1, 11, company_data_female[item])
        outSheet_T_CompanyData.write(item + 1, 12, company_data_nce[item])
        outSheet_T_CompanyData.write(item + 1, 13, company_data_nfe[item])
        outSheet_T_CompanyData.write(item + 1, 14, company_data_ce[item])
        outSheet_T_CompanyData.write(item + 1, 15, company_data_ctn[item])
        outSheet_T_CompanyData.write(item + 1, 16, company_data_proa[item])
        outSheet_T_CompanyData.write(item + 1, 17, company_data_sd[item])
        outSheet_T_CompanyData.write(item + 1, 18, company_data_proa_last[item])
        outSheet_T_CompanyData.write(item + 1, 19, company_data_sd_last[item])
    # outWorkbook.close()


def t_director_data(sheet_column_index, loop_range):
    initial_value = 0
    for i in range(loop_range):
        for item in range(len(worksheet.col_values(sheet_column_index, 2))):
            value = worksheet.col_values(20 + initial_value, 2)[item]
            if value != " ":
                try:
                    print(str(i) + " Running on t_director_data: " + str(item))
                    director_company_ids_value = worksheet.col_values(0, 2)[item]
                    director_name_kh_value = worksheet.col_values(sheet_column_index + initial_value, 2)[item]
                    director_name_en_value = worksheet.col_values(sheet_column_index + 1 + initial_value, 2)[item]
                    director_post_address_value = worksheet.col_values(sheet_column_index + 2 + initial_value, 2)[item]
                    director_telephone_value = worksheet.col_values(sheet_column_index + 3 + initial_value, 2)[item]
                    director_chairman_value = worksheet.col_values(sheet_column_index + 4 + initial_value, 2)[item]

                    director_company_ids.append(director_company_ids_value)
                    director_list.append("Director " + str(i + 1))
                    director_name_kh.append(director_name_kh_value)
                    director_name_en.append(director_name_en_value)
                    director_post_address.append(director_post_address_value)
                    director_telephone.append(director_telephone_value)
                    director_chairman.append(director_chairman_value)
                except:
                    pass
        initial_value = initial_value + 5
    for item in range(len(director_name_kh)):
        outSheet_T_Director_Data.write(item + 1, 0, director_company_ids[item])
        outSheet_T_Director_Data.write(item + 1, 1, director_list[item])
        outSheet_T_Director_Data.write(item + 1, 2, director_name_kh[item])
        outSheet_T_Director_Data.write(item + 1, 3, director_name_en[item])
        outSheet_T_Director_Data.write(item + 1, 4, director_post_address[item])
        outSheet_T_Director_Data.write(item + 1, 5, director_telephone[item])
        outSheet_T_Director_Data.write(item + 1, 6, director_chairman[item])
    outWorkbook.close()


t_business_category(45, 20)
master_company_code()
company_data()
t_director_data(20, 5)


# def copy_data():
#     # opening the source excel file
#     filename = "1_18238.xlsx"
#     wb1 = xl.load_workbook(filename)
#     ws1 = wb1.worksheets[0]
#
#     # opening the destination excel file
#     filename1 = "Data_cleaning_1_18238.xlsx"
#     wb2 = xl.load_workbook(filename1)
#     ws2 = wb2.active
#
#     # calculate total number of rows and
#     # columns in source excel file
#     mr = ws1.max_row
#     mc = ws1.max_column
#
#     # copying the cell values from source
#     # excel file to destination excel file
#     for i in range(1, mr + 1):
#         for j in range(1, mc + 1):
#             # reading cell value from source excel file
#             c = ws1.cell(row=i, column=j)
#
#             # writing the read value to destination excel file
#             ws2.cell(row=i, column=j).value = c.value
#
#         # saving the destination excel file
#     wb2.save(str(filename1))
#
#
# copy_data()
